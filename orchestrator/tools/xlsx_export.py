"""
Eksport REQUIREMENTS.md do REQUIREMENTS.xlsx.

Parsuje linijki w formacie "- [FR-001] Opis wymagania" (format, ktory
agent Analityka Wymagan ma narzucony w prompcie - patrz
prompts/system_prompts.py -> REQUIREMENTS) i zapisuje jako tabele w Excelu:
ID | Opis | Data ostatniej aktualizacji pliku.
"""
import re
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font

from config import settings

_LINE_PATTERN = re.compile(r"^-\s*\[([A-Z]+-\d+)\]\s*(.+)$")


def export_requirements_to_xlsx() -> None:
    with open(settings.REQUIREMENTS_FILE, encoding="utf-8") as f:
        lines = f.readlines()

    rows = []
    for line in lines:
        match = _LINE_PATTERN.match(line.strip())
        if match:
            rows.append((match.group(1), match.group(2)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Wymagania"

    ws.append(["ID", "Opis wymagania"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for req_id, description in rows:
        ws.append([req_id, description])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 100

    ws2 = wb.create_sheet("Info")
    ws2.append(["Wygenerowano", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws2.append(["Liczba wymagan", len(rows)])
    ws2.column_dimensions["A"].width = 20

    wb.save(settings.REQUIREMENTS_XLSX)
